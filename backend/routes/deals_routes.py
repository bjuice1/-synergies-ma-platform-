"""API routes for deal management."""
from flask import Blueprint, request, jsonify, send_file
from backend.app.models.deal import Deal
from backend.app.models.company import Company
from backend.app.models.synergy import Synergy
from backend.app.models.lever import DealLever, SynergyLever, BenchmarkDataPoint
from backend.app.services import synergy_generator  # Direct import to avoid __init__ cascade
from backend.app.extensions import db
from backend.utils.auth_decorators import require_role
from backend.utils.exceptions import ValidationError, NotFoundError
from datetime import datetime
from sqlalchemy import func
import logging

logger = logging.getLogger(__name__)

LEVER_ACTIVITY_TEMPLATES: dict = {
    'IT': [
        ('cost_reduction', 'Application rationalisation & decommissioning'),
        ('cost_reduction', 'Infrastructure & data centre consolidation'),
        ('cost_reduction', 'IT vendor contract renegotiation'),
    ],
    'Finance': [
        ('cost_reduction', 'Finance function consolidation to shared services'),
        ('cost_reduction', 'ERP & reporting systems harmonisation'),
        ('cost_reduction', 'External audit & advisory fee reduction'),
    ],
    'HR': [
        ('cost_reduction', 'Redundant role elimination & org restructuring'),
        ('cost_reduction', 'Benefits & compensation plan harmonisation'),
        ('cost_reduction', 'HRIS, payroll & LMS systems consolidation'),
    ],
    'Operations': [
        ('cost_reduction', 'Operations & manufacturing footprint rationalisation'),
        ('cost_reduction', 'Supply chain & logistics network optimisation'),
        ('cost_reduction', 'Quality & compliance systems harmonisation'),
    ],
    'Procurement': [
        ('cost_reduction', 'Vendor consolidation & volume leverage renegotiation'),
        ('cost_reduction', 'Indirect spend category management'),
        ('cost_reduction', 'Procurement function & systems integration'),
    ],
    'Real Estate': [
        ('cost_reduction', 'Facility footprint reduction & lease exits'),
        ('cost_reduction', 'Headquarters & office space co-location'),
        ('cost_reduction', 'Property management & utilities optimisation'),
    ],
    'Revenue': [
        ('revenue_enhancement', 'Cross-sell acquirer products to target customer base'),
        ('revenue_enhancement', 'Geographic expansion via acquired market presence'),
        ('revenue_enhancement', 'Bundled product/service premium pricing opportunity'),
    ],
}


def auto_generate_deal_levers(deal, acquirer, target):
    """
    Auto-generate DealLever rows for a new deal using benchmark averages.

    Uses all BenchmarkDataPoints to compute pct_low = min, pct_high = max,
    pct_median = mean for each lever. calculated_value = combined_revenue * pct / 100.

    Also creates 3 templated Synergy activity records per lever so analysts
    immediately see specific ideas rather than an empty card.

    Revenue lever is skipped if no datapoints exist.
    Never raises — lever gen failure must not break deal creation.
    """
    try:
        combined_revenue = 0
        if acquirer and acquirer.revenue_usd:
            combined_revenue += acquirer.revenue_usd
        if target and target.revenue_usd:
            combined_revenue += target.revenue_usd

        levers = SynergyLever.query.order_by(SynergyLever.sort_order).all()
        benchmark_n_result = db.session.query(
            func.count(db.distinct(BenchmarkDataPoint.project_id))
        ).scalar() or 0

        created = 0
        for lever in levers:
            existing = DealLever.query.filter_by(deal_id=deal.id, lever_id=lever.id).first()
            if existing:
                continue

            pct_values = [
                row[0] for row in
                db.session.query(BenchmarkDataPoint.synergy_pct)
                .filter(BenchmarkDataPoint.lever_id == lever.id)
                .all()
            ]

            if not pct_values:
                continue

            pct_sorted = sorted(pct_values)
            pct_low = round(pct_sorted[0], 2)
            pct_high = round(pct_sorted[-1], 2)
            pct_median = round(sum(pct_sorted) / len(pct_sorted), 2)

            val_low  = int(combined_revenue * pct_low  / 100) if combined_revenue else None
            val_high = int(combined_revenue * pct_high / 100) if combined_revenue else None

            dl = DealLever(
                deal_id=deal.id,
                lever_id=lever.id,
                benchmark_pct_low=pct_low,
                benchmark_pct_high=pct_high,
                benchmark_pct_median=pct_median,
                benchmark_n=benchmark_n_result,
                combined_baseline_usd=None,
                calculated_value_low=val_low,
                calculated_value_high=val_high,
                status='identified',
                confidence='medium',
                advisor_notes=None,
            )
            db.session.add(dl)
            db.session.flush()  # get dl.id before creating activities

            # Create 3 templated synergy activities under this lever
            templates = LEVER_ACTIVITY_TEMPLATES.get(lever.name, [])
            n = len(templates) or 1
            for synergy_type, description in templates:
                act = Synergy(
                    company1_id=acquirer.id,
                    company2_id=target.id,
                    deal_id=deal.id,
                    deal_lever_id=dl.id,
                    synergy_type=synergy_type,
                    description=description,
                    value_low=int(val_low / n) if val_low else None,
                    value_high=int(val_high / n) if val_high else None,
                    confidence_score=60,
                    status='identified',
                )
                db.session.add(act)

            created += 1

        if created > 0:
            db.session.flush()
            logger.info(f"Auto-generated {created} DealLevers for deal {deal.id}")

    except Exception as e:
        logger.error(f"Failed to auto-generate levers for deal {deal.id}: {e}", exc_info=True)


bp = Blueprint('deals', __name__, url_prefix='/api/deals')


@bp.route('', methods=['GET'])
@require_role('viewer', 'analyst', 'admin')
def get_deals():
    """
    Get all deals with optional filtering.

    Query Parameters:
        status: Filter by status (draft, active, closed, cancelled)
        acquirer_id: Filter by acquiring company
        target_id: Filter by target company
    """
    try:
        # Get query parameters
        status = request.args.get('status')
        acquirer_id = request.args.get('acquirer_id', type=int)
        target_id = request.args.get('target_id', type=int)

        # Build query
        query = Deal.query

        if status:
            query = query.filter(Deal.status == status)
        if acquirer_id:
            query = query.filter(Deal.acquirer_id == acquirer_id)
        if target_id:
            query = query.filter(Deal.target_id == target_id)

        # Order by most recent first
        deals = query.order_by(Deal.created_at.desc()).all()

        return jsonify([deal.to_dict(include_synergies=True) for deal in deals]), 200

    except Exception as e:
        logger.error(f"Error fetching deals: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>', methods=['GET'])
@require_role('viewer', 'analyst', 'admin')
def get_deal(deal_id):
    """Get a specific deal by ID."""
    try:
        deal = Deal.query.get_or_404(deal_id)
        return jsonify(deal.to_dict(include_synergies=True)), 200
    except NotFoundError as e:
        return jsonify({'error': 'Deal not found'}), 404
    except Exception as e:
        logger.error(f"Error fetching deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('', methods=['POST'])
@require_role('analyst', 'admin')
def create_deal():
    """
    Create a new deal.

    Request Body:
        name: Deal name (required)
        deal_type: Type of deal (acquisition, merger, jv, partnership)
        deal_size_usd: Deal size in USD
        close_date: Expected close date (ISO format)
        strategic_rationale: Why this deal makes sense
        acquirer: Acquirer company details (new or existing)
        target: Target company details (new or existing)
    """
    try:
        data = request.get_json()

        # Validate required fields
        if not data.get('name'):
            return jsonify({'error': 'Deal name is required'}), 400
        if not data.get('acquirer'):
            return jsonify({'error': 'Acquirer company details are required'}), 400
        if not data.get('target'):
            return jsonify({'error': 'Target company details are required'}), 400

        # Create or get acquirer company
        acquirer_data = data['acquirer']
        if acquirer_data.get('id'):
            acquirer = Company.query.get(acquirer_data['id'])
            if not acquirer:
                return jsonify({'error': 'Acquirer company not found'}), 404
        else:
            # Create new acquirer company
            acquirer = Company(
                name=acquirer_data.get('name'),
                industry=acquirer_data.get('industry', 'Unknown'),
                description=acquirer_data.get('description'),
                revenue_usd=acquirer_data.get('revenue_usd'),
                employees=acquirer_data.get('employees'),
                geography=acquirer_data.get('geography', []),
                products=acquirer_data.get('products', []),
                tech_stack=acquirer_data.get('tech_stack', []),
                strengths=acquirer_data.get('strengths', []),
                weaknesses=acquirer_data.get('weaknesses', [])
            )
            db.session.add(acquirer)
            db.session.flush()  # Get acquirer ID

        # Create or get target company
        target_data = data['target']
        if target_data.get('id'):
            target = Company.query.get(target_data['id'])
            if not target:
                return jsonify({'error': 'Target company not found'}), 404
        else:
            # Create new target company
            target = Company(
                name=target_data.get('name'),
                industry=target_data.get('industry', 'Unknown'),
                description=target_data.get('description'),
                revenue_usd=target_data.get('revenue_usd'),
                employees=target_data.get('employees'),
                geography=target_data.get('geography', []),
                products=target_data.get('products', []),
                tech_stack=target_data.get('tech_stack', []),
                strengths=target_data.get('strengths', []),
                weaknesses=target_data.get('weaknesses', [])
            )
            db.session.add(target)
            db.session.flush()  # Get target ID

        # Parse close_date if provided
        close_date = None
        if data.get('close_date'):
            try:
                close_date = datetime.fromisoformat(data['close_date'].replace('Z', '+00:00')).date()
            except ValueError:
                return jsonify({'error': 'Invalid close_date format. Use ISO 8601 (YYYY-MM-DD)'}), 400

        # Create deal
        deal = Deal(
            name=data['name'],
            deal_type=data.get('deal_type', 'acquisition'),
            deal_size_usd=data.get('deal_size_usd'),
            close_date=close_date,
            strategic_rationale=data.get('strategic_rationale'),
            acquirer_id=acquirer.id,
            target_id=target.id,
            deal_briefing_document=data.get('deal_briefing_document'),
            status='draft'
        )

        db.session.add(deal)
        db.session.flush()  # Get deal ID before lever generation
        auto_generate_deal_levers(deal, acquirer, target)
        db.session.commit()

        return jsonify(deal.to_dict()), 201

    except ValidationError as e:
        db.session.rollback()
        logger.warning(f"Validation error creating deal: {e}")
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating deal: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>', methods=['PUT'])
@require_role('analyst', 'admin')
def update_deal(deal_id):
    """Update an existing deal."""
    try:
        deal = Deal.query.get_or_404(deal_id)
        data = request.get_json()

        # Update fields
        if 'name' in data:
            deal.name = data['name']
        if 'deal_type' in data:
            deal.deal_type = data['deal_type']
        if 'deal_size_usd' in data:
            deal.deal_size_usd = data['deal_size_usd']
        if 'close_date' in data:
            if data['close_date']:
                deal.close_date = datetime.fromisoformat(data['close_date'].replace('Z', '+00:00')).date()
            else:
                deal.close_date = None
        if 'strategic_rationale' in data:
            deal.strategic_rationale = data['strategic_rationale']
        if 'status' in data:
            deal.status = data['status']
        if 'deal_briefing_document' in data:
            deal.deal_briefing_document = data['deal_briefing_document']

        db.session.commit()

        return jsonify(deal.to_dict(include_synergies=True)), 200

    except ValidationError as e:
        db.session.rollback()
        logger.warning(f"Validation error updating deal {deal_id}: {e}")
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>', methods=['DELETE'])
@require_role('admin')
def delete_deal(deal_id):
    """Delete a deal and all associated synergies."""
    try:
        deal = Deal.query.get_or_404(deal_id)
        db.session.delete(deal)
        db.session.commit()

        return jsonify({'message': 'Deal deleted successfully'}), 200

    except NotFoundError as e:
        return jsonify({'error': 'Deal not found'}), 404
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/levers', methods=['GET'])
@require_role('viewer', 'analyst', 'admin')
def get_deal_levers(deal_id):
    """
    Get all synergy levers for a deal, including benchmark data,
    cost baselines, calculated opportunity ranges, and activities.
    This is the primary view for deal synergy analysis.
    """
    try:
        deal = Deal.query.get_or_404(deal_id)
        deal_levers = (
            DealLever.query
            .filter_by(deal_id=deal_id)
            .join(SynergyLever)
            .order_by(SynergyLever.sort_order)
            .all()
        )

        result = []
        for dl in deal_levers:
            lever_data = dl.to_dict()
            # Include activities (specific synergy items under this lever)
            activities = Synergy.query.filter_by(deal_lever_id=dl.id).all()
            lever_data['activities'] = [
                {
                    'id': a.id,
                    'synergy_type': a.synergy_type,
                    'description': a.description,
                    'value_low': a.value_low,
                    'value_high': a.value_high,
                    'status': a.status,
                    'confidence_score': a.confidence_score,
                }
                for a in activities
            ]
            result.append(lever_data)

        # Summary totals
        cost_levers = [dl for dl in deal_levers if dl.lever and dl.lever.lever_type == 'cost']
        total_low = sum(dl.calculated_value_low or 0 for dl in cost_levers)
        total_high = sum(dl.calculated_value_high or 0 for dl in cost_levers)
        combined_revenue = (deal.acquirer.revenue_usd or 0) + (deal.target.revenue_usd or 0) if deal.acquirer and deal.target else 0

        return jsonify({
            'deal_id': deal_id,
            'levers': result,
            'summary': {
                'total_cost_synergy_low': total_low,
                'total_cost_synergy_high': total_high,
                'combined_revenue': combined_revenue,
                'total_pct_low': round(total_low / combined_revenue * 100, 1) if combined_revenue else None,
                'total_pct_high': round(total_high / combined_revenue * 100, 1) if combined_revenue else None,
                'benchmark_n': deal_levers[0].benchmark_n if deal_levers else 0,
            }
        }), 200

    except Exception as e:
        logger.error(f"Error fetching levers for deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/companies/<int:company_id>', methods=['PATCH'])
@require_role('analyst', 'admin')
def update_company(company_id):
    """Update a company's financial fields (revenue, employees)."""
    try:
        company = Company.query.get_or_404(company_id)
        data = request.get_json() or {}
        if 'revenue_usd' in data:
            company.revenue_usd = int(data['revenue_usd']) if data['revenue_usd'] is not None else None
        if 'employees' in data:
            company.employees = int(data['employees']) if data['employees'] is not None else None
        if 'name' in data and data['name']:
            company.name = data['name'].strip()
        db.session.commit()
        return jsonify(company.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating company {company_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/levers/<int:lever_id>', methods=['PATCH'])
@require_role('analyst', 'admin')
def update_deal_lever(deal_id, lever_id):
    """Update a DealLever's analyst notes, status, or confidence."""
    try:
        dl = DealLever.query.filter_by(deal_id=deal_id, id=lever_id).first_or_404()
        data = request.get_json() or {}

        if 'advisor_notes' in data:
            dl.advisor_notes = data['advisor_notes'] or None
        if 'status' in data and data['status'] in ('identified', 'in_analysis', 'validated', 'excluded'):
            dl.status = data['status']
        if 'confidence' in data and data['confidence'] in ('high', 'medium', 'low'):
            dl.confidence = data['confidence']
        if 'environment_data' in data and isinstance(data['environment_data'], dict):
            dl.environment_data = data['environment_data']
        if 'assigned_to_id' in data:
            dl.assigned_to_id = data['assigned_to_id'] or None

        db.session.commit()
        return jsonify(dl.to_dict()), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating lever {lever_id} for deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/levers/<int:lever_id>/comments', methods=['GET'])
@require_role('analyst', 'admin')
def get_lever_comments(deal_id, lever_id):
    """List comments for a deal lever."""
    try:
        from backend.app.models.lever import LeverComment
        dl = DealLever.query.filter_by(deal_id=deal_id, id=lever_id).first_or_404()
        comments = LeverComment.query.filter_by(deal_lever_id=dl.id).order_by(LeverComment.created_at).all()
        return jsonify([c.to_dict() for c in comments]), 200
    except Exception as e:
        logger.error(f"Error fetching comments for lever {lever_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/levers/<int:lever_id>/comments', methods=['POST'])
@require_role('analyst', 'admin')
def post_lever_comment(deal_id, lever_id):
    """Post a comment on a deal lever."""
    try:
        from backend.app.models.lever import LeverComment
        from flask_jwt_extended import get_jwt_identity
        dl = DealLever.query.filter_by(deal_id=deal_id, id=lever_id).first_or_404()
        data = request.get_json() or {}
        body = (data.get('body') or '').strip()
        if not body:
            return jsonify({'error': 'Comment body is required'}), 400
        user_id = int(get_jwt_identity())
        comment = LeverComment(deal_lever_id=dl.id, user_id=user_id, body=body)
        db.session.add(comment)
        db.session.commit()
        return jsonify(comment.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error posting comment for lever {lever_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/levers/<int:lever_id>/refine', methods=['POST'])
@require_role('analyst', 'admin')
def refine_lever_estimate(deal_id, lever_id):
    """
    Send filled diligence Q&A + deal context to Claude.
    Claude returns a deal-specific pct_low/pct_high (within benchmark bounds) + rationale.
    """
    try:
        import os, json
        from anthropic import Anthropic

        dl = DealLever.query.filter_by(deal_id=deal_id, id=lever_id).first_or_404()
        deal = Deal.query.get_or_404(deal_id)

        qa = dl.environment_data or {}
        filled_qa = {q: a for q, a in qa.items() if a and a.strip()}
        if not filled_qa:
            return jsonify({'error': 'No diligence answers found. Fill in some questions first.'}), 400

        benchmark_low  = dl.benchmark_pct_low  or 0
        benchmark_high = dl.benchmark_pct_high or 0
        lever_name     = dl.lever.name if dl.lever else 'this lever'

        acq_rev  = deal.acquirer.revenue_usd  if deal.acquirer else None
        tgt_rev  = deal.target.revenue_usd   if deal.target  else None
        combined = (acq_rev or 0) + (tgt_rev or 0)
        acq_name = deal.acquirer.name if deal.acquirer else 'Acquirer'
        tgt_name = deal.target.name  if deal.target  else 'Target'

        qa_text = '\n'.join(f'Q: {q}\nA: {a}' for q, a in filled_qa.items())

        prompt = f"""You are a senior M&A synergy analyst. Your task is to estimate the {lever_name} synergy opportunity for a specific deal, given benchmark data and diligence findings.

## Deal Context
- Transaction: {acq_name} acquires {tgt_name}
- Combined revenue: ${combined/1e6:.0f}M
- Lever: {lever_name}
- Benchmark range (comparable deals): {benchmark_low:.2f}%–{benchmark_high:.2f}% of combined revenue

## Diligence Findings
{qa_text}

## Your Task
Based on the diligence findings, position this deal within (or explain deviation from) the benchmark range.
- Consider factors that suggest higher or lower synergy potential
- Stay within {benchmark_low:.2f}%–{benchmark_high:.2f}% unless the evidence strongly justifies deviation
- Be specific about which findings drove your estimate

Respond ONLY with a JSON object in this exact format (no markdown, no explanation outside JSON):
{{
  "pct_low": <float between 0 and 10>,
  "pct_high": <float between 0 and 10>,
  "rationale": "<2-4 sentence explanation of positioning within the range>"
}}"""

        client = Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
        message = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=512,
            messages=[{'role': 'user', 'content': prompt}]
        )
        raw = message.content[0].text.strip()

        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            import re
            m = re.search(r'\{.*\}', raw, re.DOTALL)
            if not m:
                return jsonify({'error': 'Claude returned unexpected format', 'raw': raw}), 500
            result = json.loads(m.group())

        pct_low  = float(result.get('pct_low',  benchmark_low))
        pct_high = float(result.get('pct_high', benchmark_high))
        rationale = str(result.get('rationale', ''))

        dl.refined_pct_low         = pct_low
        dl.refined_pct_high        = pct_high
        dl.refinement_rationale    = rationale

        # Recalculate $ values using refined pcts × combined revenue
        if combined > 0:
            dl.calculated_value_low  = int(combined * pct_low  / 100)
            dl.calculated_value_high = int(combined * pct_high / 100)

        db.session.commit()
        return jsonify(dl.to_dict()), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error refining lever {lever_id} for deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/populate-from-brief', methods=['POST'])
@require_role('analyst', 'admin')
def populate_from_brief(deal_id):
    """
    Read the stored deal briefing document and use Claude to pre-populate
    environment_data (diligence Q&A) for every lever that has playbook questions.
    Only fills blanks — never overwrites analyst work.
    """
    try:
        from anthropic import Anthropic
        from backend.app.models.playbook import LeverPlaybook
        import json as _json

        deal = Deal.query.get_or_404(deal_id)
        if not deal.deal_briefing_document:
            return jsonify({'error': 'No briefing document stored for this deal. Upload one first.'}), 400

        deal_levers = (
            DealLever.query
            .filter_by(deal_id=deal_id)
            .join(SynergyLever)
            .order_by(SynergyLever.sort_order)
            .all()
        )

        # Build questions block for the prompt
        lever_questions: dict[str, list[str]] = {}
        for dl in deal_levers:
            if not dl.lever:
                continue
            playbook = LeverPlaybook.query.filter_by(lever_id=dl.lever.id).first()
            if playbook and playbook.diligence_questions:
                lever_questions[dl.lever.name] = playbook.diligence_questions

        if not lever_questions:
            return jsonify({'message': 'No diligence questions configured in playbooks', 'updated': 0}), 200

        questions_block = "\n\n".join(
            f"## {lever}\n" + "\n".join(f"- {q}" for q in qs)
            for lever, qs in lever_questions.items()
        )

        client = Anthropic()
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": f"""You are an M&A integration analyst. Read the deal brief below and answer the diligence questions for each lever where the brief contains relevant information.

DEAL BRIEF:
{deal.deal_briefing_document}

DILIGENCE QUESTIONS BY LEVER:
{questions_block}

Instructions:
- Only answer questions where the brief contains explicit or strongly implied information
- Keep answers concise (1-2 sentences), factual, and grounded in the brief
- If the brief does not address a question, skip that question entirely
- Do not invent information not present in the brief

Return ONLY a JSON object. No markdown, no explanation. Format:
{{"IT": {{"question text": "answer text"}}, "Finance": {{"question text": "answer text"}}, ...}}"""
            }]
        )

        raw = message.content[0].text.strip()
        # Strip markdown code fences if present
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()

        try:
            answers = _json.loads(raw)
        except _json.JSONDecodeError:
            logger.error(f"Claude returned invalid JSON for deal {deal_id}: {raw[:200]}")
            return jsonify({'error': 'AI returned malformed data — try again'}), 500

        # Apply answers, never overwriting existing analyst entries
        updated = 0
        for dl in deal_levers:
            if not dl.lever:
                continue
            lever_answers = answers.get(dl.lever.name, {})
            if not lever_answers:
                continue
            existing = dl.environment_data or {}
            merged = dict(existing)
            changed = False
            for q, a in lever_answers.items():
                if not merged.get(q) and a and len(a.strip()) > 5:
                    merged[q] = a.strip()
                    changed = True
            if changed:
                dl.environment_data = merged
                updated += 1

        db.session.commit()

        # Return refreshed lever list
        refreshed = (
            DealLever.query.filter_by(deal_id=deal_id)
            .join(SynergyLever).order_by(SynergyLever.sort_order).all()
        )
        result = []
        for dl in refreshed:
            lever_data = dl.to_dict()
            activities = Synergy.query.filter_by(deal_lever_id=dl.id).all()
            lever_data['activities'] = [
                {'id': a.id, 'synergy_type': a.synergy_type, 'description': a.description,
                 'value_low': a.value_low, 'value_high': a.value_high,
                 'status': a.status, 'confidence_score': a.confidence_score}
                for a in activities
            ]
            result.append(lever_data)

        return jsonify({'updated': updated, 'levers': result}), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error populating brief for deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/generate-synergies', methods=['POST'])
@require_role('analyst', 'admin')
def generate_synergies(deal_id):
    """
    Generate synergies for a deal using AI/rules.

    This endpoint analyzes the deal context and creates potential synergies.
    Phase 1: Simple rule-based patterns.
    Future: LLM-powered analysis.
    """
    try:
        deal = Deal.query.get_or_404(deal_id)

        # Check if synergies already exist (idempotency)
        existing_count = deal.synergies.count()
        if existing_count > 0:
            return jsonify({
                'message': f'Deal already has {existing_count} synergies. Delete them first to regenerate.',
                'synergies': [s.to_dict() for s in deal.synergies.all()]
            }), 200

        # Get acquirer and target companies
        acquirer = deal.acquirer
        target = deal.target

        if not acquirer or not target:
            return jsonify({'error': 'Deal must have both acquirer and target companies'}), 400

        # Generate synergies using service
        synergy_dicts = synergy_generator.generate_synergies_for_deal(deal, acquirer, target)

        # Create synergy records
        created_synergies = []
        for synergy_data in synergy_dicts:
            synergy = Synergy(**synergy_data)
            db.session.add(synergy)
            created_synergies.append(synergy)

        db.session.commit()

        return jsonify({
            'message': f'Generated {len(created_synergies)} synergies',
            'synergies': [s.to_dict() for s in created_synergies]
        }), 201

    except ValidationError as e:
        db.session.rollback()
        logger.warning(f"Validation error generating synergies for deal {deal_id}: {e}")
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error generating synergies for deal {deal_id}: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@bp.route('/<int:deal_id>/export/excel', methods=['GET'])
@require_role('analyst', 'admin')
def export_excel(deal_id):
    """Export deal lever analysis as Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        deal = Deal.query.get_or_404(deal_id)
        levers = DealLever.query.filter_by(deal_id=deal_id).join(SynergyLever).order_by(SynergyLever.sort_order).all()

        wb = openpyxl.Workbook()

        # ── Colours ──────────────────────────────────────────────
        ORANGE  = 'FFD04A02'
        WHITE   = 'FFFFFFFF'
        LGRAY   = 'FFF5F5F5'
        MGRAY   = 'FFE5E7EB'
        DKTEXT  = 'FF111827'
        SUBTEXT = 'FF6B7280'

        def hdr(ws, row, col, value, bold=True, bg=ORANGE, fg=WHITE, size=11):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color=fg, size=size)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            return c

        def cell(ws, row, col, value, bold=False, bg=None, fg=DKTEXT, size=10, wrap=False, align='left'):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=bold, color=fg, size=size)
            if bg:
                c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
            return c

        thin = Border(
            left=Side(style='thin', color=MGRAY),
            right=Side(style='thin', color=MGRAY),
            top=Side(style='thin', color=MGRAY),
            bottom=Side(style='thin', color=MGRAY),
        )

        def fmt_m(val):
            if val is None: return '—'
            if val >= 1_000_000: return f'${val/1_000_000:.1f}M'
            if val >= 1_000: return f'${val/1_000:.0f}K'
            return f'${val:,.0f}'

        # ── Sheet 1: Summary ─────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Summary'
        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 20
        ws1.row_dimensions[1].height = 36

        # Title
        ws1.merge_cells('A1:E1')
        t = ws1['A1']
        t.value = deal.name
        t.font = Font(bold=True, size=16, color=WHITE)
        t.fill = PatternFill('solid', fgColor=ORANGE)
        t.alignment = Alignment(horizontal='left', vertical='center')

        ws1.merge_cells('A2:E2')
        sub = ws1['A2']
        acq = deal.acquirer.name if deal.acquirer else ''
        tgt = deal.target.name if deal.target else ''
        sub.value = f'{acq}  →  {tgt}   |   {deal.deal_type.capitalize()}   |   Generated {datetime.utcnow().strftime("%d %b %Y")}'
        sub.font = Font(size=10, color=SUBTEXT)
        sub.fill = PatternFill('solid', fgColor=LGRAY)
        sub.alignment = Alignment(horizontal='left', vertical='center')
        ws1.row_dimensions[2].height = 22

        # Deal metrics
        r = 4
        metrics = [
            ('Deal Size', fmt_m(deal.deal_size_usd)),
            ('Acquirer Revenue', fmt_m(deal.acquirer.revenue_usd if deal.acquirer else None)),
            ('Target Revenue', fmt_m(deal.target.revenue_usd if deal.target else None)),
            ('Acquirer Employees', f'{deal.acquirer.employees:,}' if deal.acquirer and deal.acquirer.employees else '—'),
            ('Target Employees', f'{deal.target.employees:,}' if deal.target and deal.target.employees else '—'),
        ]
        for label, val in metrics:
            cell(ws1, r, 1, label, bold=True, fg=SUBTEXT, size=9)
            cell(ws1, r, 2, val, bold=True, size=11)
            r += 1

        # Synergy summary header
        r += 1
        cost_levers = [l for l in levers if l.lever.lever_type == 'cost']
        rev_levers  = [l for l in levers if l.lever.lever_type == 'revenue']
        combined_rev = (deal.acquirer.revenue_usd or 0) + (deal.target.revenue_usd or 0)
        total_low  = sum((l.calculated_value_low  or 0) for l in cost_levers)
        total_high = sum((l.calculated_value_high or 0) for l in cost_levers)

        hdr(ws1, r, 1, 'Synergy Opportunity', size=12)
        ws1.merge_cells(f'A{r}:E{r}')
        r += 1

        headers = ['', 'Low', 'High', '% Rev (Low)', '% Rev (High)']
        for ci, h in enumerate(headers, 1):
            hdr(ws1, r, ci, h, size=9, bg=MGRAY, fg=DKTEXT)
        r += 1

        def pct(val):
            if not combined_rev or val is None: return '—'
            return f'{val/combined_rev*100:.1f}%'

        for lever in cost_levers:
            bg = LGRAY if r % 2 == 0 else WHITE
            cell(ws1, r, 1, lever.lever.name, bold=True, bg=bg)
            cell(ws1, r, 2, fmt_m(lever.calculated_value_low), bg=bg, align='right')
            cell(ws1, r, 3, fmt_m(lever.calculated_value_high), bg=bg, align='right')
            cell(ws1, r, 4, pct(lever.calculated_value_low), bg=bg, align='right', fg=SUBTEXT)
            cell(ws1, r, 5, pct(lever.calculated_value_high), bg=bg, align='right', fg=SUBTEXT)
            r += 1

        # Total row
        cell(ws1, r, 1, 'Total Cost Synergy', bold=True, bg=LGRAY)
        cell(ws1, r, 2, fmt_m(total_low),  bold=True, bg=LGRAY, align='right')
        cell(ws1, r, 3, fmt_m(total_high), bold=True, bg=LGRAY, align='right')
        cell(ws1, r, 4, pct(total_low),  bold=True, bg=LGRAY, align='right', fg=SUBTEXT)
        cell(ws1, r, 5, pct(total_high), bold=True, bg=LGRAY, align='right', fg=SUBTEXT)
        r += 1

        for lever in rev_levers:
            bg = LGRAY if r % 2 == 0 else WHITE
            cell(ws1, r, 1, f'{lever.lever.name} (Revenue)', bold=True, bg=bg, fg='FF0369A1')
            cell(ws1, r, 2, fmt_m(lever.calculated_value_low), bg=bg, align='right')
            cell(ws1, r, 3, fmt_m(lever.calculated_value_high), bg=bg, align='right')
            cell(ws1, r, 4, pct(lever.calculated_value_low), bg=bg, align='right', fg=SUBTEXT)
            cell(ws1, r, 5, pct(lever.calculated_value_high), bg=bg, align='right', fg=SUBTEXT)
            r += 1

        if deal.strategic_rationale:
            r += 1
            hdr(ws1, r, 1, 'Strategic Rationale', size=10, bg=MGRAY, fg=DKTEXT)
            ws1.merge_cells(f'A{r}:E{r}')
            r += 1
            ws1.merge_cells(f'A{r}:E{r+2}')
            c = ws1.cell(row=r, column=1, value=deal.strategic_rationale)
            c.font = Font(size=10, color=DKTEXT)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws1.row_dimensions[r].height = 54

        # ── Sheet 2: Lever Detail ─────────────────────────────────
        ws2 = wb.create_sheet('Lever Detail')
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 16
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 14
        ws2.column_dimensions['G'].width = 40

        ws2.merge_cells('A1:G1')
        t2 = ws2['A1']
        t2.value = f'{deal.name} — Lever Detail'
        t2.font = Font(bold=True, size=14, color=WHITE)
        t2.fill = PatternFill('solid', fgColor=ORANGE)
        t2.alignment = Alignment(horizontal='left', vertical='center')
        ws2.row_dimensions[1].height = 32

        hdrs2 = ['Lever', 'Synergy Low', 'Synergy High', 'Bmark Low%', 'Bmark High%', 'Confidence', 'Advisor Notes']
        r2 = 2
        for ci, h in enumerate(hdrs2, 1):
            hdr(ws2, r2, ci, h, size=9, bg=MGRAY, fg=DKTEXT)
        r2 += 1

        for lever in levers:
            bg = LGRAY if r2 % 2 == 0 else WHITE
            cell(ws2, r2, 1, lever.lever.name, bold=True, bg=bg)
            cell(ws2, r2, 2, fmt_m(lever.calculated_value_low), bg=bg, align='right')
            cell(ws2, r2, 3, fmt_m(lever.calculated_value_high), bg=bg, align='right')
            cell(ws2, r2, 4, f'{lever.benchmark_pct_low:.2f}%' if lever.benchmark_pct_low is not None else '—', bg=bg, align='right', fg=SUBTEXT)
            cell(ws2, r2, 5, f'{lever.benchmark_pct_high:.2f}%' if lever.benchmark_pct_high is not None else '—', bg=bg, align='right', fg=SUBTEXT)
            cell(ws2, r2, 6, lever.confidence or 'medium', bg=bg, fg=SUBTEXT)
            cell(ws2, r2, 7, lever.advisor_notes or '', bg=bg, wrap=True)
            r2 += 1

        # ── Sheet 3: Activities ────────────────────────────────────
        ws3 = wb.create_sheet('Activities')
        ws3.column_dimensions['A'].width = 18
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 44
        ws3.column_dimensions['D'].width = 14
        ws3.column_dimensions['E'].width = 14
        ws3.column_dimensions['F'].width = 12

        ws3.merge_cells('A1:F1')
        t3 = ws3['A1']
        t3.value = f'{deal.name} — Integration Activities'
        t3.font = Font(bold=True, size=14, color=WHITE)
        t3.fill = PatternFill('solid', fgColor=ORANGE)
        t3.alignment = Alignment(horizontal='left', vertical='center')
        ws3.row_dimensions[1].height = 32

        hdrs3 = ['Lever', 'Type', 'Description', 'Value Low', 'Value High', 'Status']
        r3 = 2
        for ci, h in enumerate(hdrs3, 1):
            hdr(ws3, r3, ci, h, size=9, bg=MGRAY, fg=DKTEXT)
        r3 += 1

        for lever in levers:
            activities = Synergy.query.filter_by(deal_id=deal_id, deal_lever_id=lever.id).all()
            for act in activities:
                bg = LGRAY if r3 % 2 == 0 else WHITE
                cell(ws3, r3, 1, lever.lever.name, bold=True, bg=bg)
                cell(ws3, r3, 2, act.synergy_type.replace('_', ' ').title(), bg=bg, fg=SUBTEXT)
                cell(ws3, r3, 3, act.description or '', bg=bg, wrap=True)
                cell(ws3, r3, 4, fmt_m(act.value_low), bg=bg, align='right')
                cell(ws3, r3, 5, fmt_m(act.value_high), bg=bg, align='right')
                cell(ws3, r3, 6, act.status or 'identified', bg=bg, fg=SUBTEXT)
                r3 += 1

        # ── Stream response ───────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = deal.name.replace(' ', '_').replace('/', '-')[:60]
        filename = f'{safe_name}_synergy_analysis.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        logger.error(f'Excel export error for deal {deal_id}: {e}', exc_info=True)
        return jsonify({'error': 'Export failed'}), 500
